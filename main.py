# import os
# import openpyxl
# Excel_path="\\excel\\xlsx"
# save_path='excel'
# def create_total_xlsx():
#     if not os.path.exists(save_path+'\\'+'总表.xlsx'):
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.title='Sheet1'
#     else:
#         os.remove(save_path+'\\'+'总表.xlsx')
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.title='Sheet1'
#     for i in range(1,18):
#         data = [[i,'星期一' , '星期二','星期三','星期四','星期五','星期六','星期日']]
#         for row in data:
#             sheet.append(row)
#         for j in range(0,5):
#             data = [[i]]
#             for row in data:
#                 sheet.append(row)
#     workbook.save(save_path + '\\' + '总表.xlsx')
#
# def find_letter_position(letter):
#     letter_positions = {}
#     for index, char in enumerate(letter):
#         # 检查是否为字母
#         if char.isalpha():
#             # 如果是字母，记录其位置
#             # letter_positions[char] = letter_positions.get(char, []) + [index + 1]  # 索引从1开始
#             return index
#
#
# def add_data():
#     retval = os.getcwd()
#     retval = retval + Excel_path
#     all_files = os.listdir(retval)
#
#     save_workbook=openpyxl.load_workbook(os.getcwd()+'\\'+save_path+'\\'+'总表.xlsx')
#     save_sheet = save_workbook['Sheet1']
#     # 仅保留Excel文件
#     excel_files = [file for file in all_files if file.endswith('.xlsx') or file.endswith('.xls')]
#     for excel_file in excel_files:
#         name=excel_file.split('.')[0]
#         print(name)
#         file_path = os.path.join(retval, excel_file)
#         workbook = openpyxl.load_workbook(file_path)
#         sheet = workbook['Sheet1']
#         if sheet.cell(row=1,column=6).value == '星期一':
#             for i in range(6,13):
#                 for j in range(2, sheet.max_row+1):
#                     if not sheet.cell(row=j, column=i).value=='':
#                         class_information = sheet.cell(row=j, column=i).value
#                         if class_information == None:
#                             continue
#                         if len(class_information)>14:
#                             continue
#                         letter_index=find_letter_position(class_information)
#                         if letter_index==None:
#                             continue
#                         class_time=class_information[:letter_index]
#                         class_cycle=class_information[letter_index+4:]
#                         if class_time == "0102":
#                             save_row = 1
#                         elif class_time == "0304":
#                             save_row = 2
#                         elif class_time == "0506":
#                             save_row = 3
#                         elif class_time == "0708":
#                             save_row = 4
#                         elif class_time == "0910":
#                             save_row = 5
#                         else:
#                             continue
#                         parts =class_cycle.split('-')
#                         head=parts[0]
#                         end=parts[-1]
#                         if head==end:
#                             continue
#                         for cycle in range(int(head)-1,int(end)):
#                             old=save_sheet.cell(row=save_row + 1+cycle*6, column=i - 5 + 1).value
#                             if old == None:
#                                 old=''
#                             new=name+" 忙碌\n"
#
#                             combined_content =f"{old} {new}"
#                             save_sheet.cell(save_row + 1+cycle*6, i - 5 + 1, combined_content)
#     save_workbook.save(save_path + '\\' + '总表.xlsx')
# def main(name):
#     create_total_xlsx()
#     add_data()
#
# if __name__ == '__main__':
#     main('PyCharm')


#                                新版教务
#   使用注意事项：所有课程信息需要按照0102B123这个格式，
#   如果是实训或实验的课程需要改成12B123(只要是12开头加一个大写字母就可以后面只要有就行)
#   跑脚本的时候不能打开xlsx文件
import os
import openpyxl
import datetime
import xlrd
import pandas as pd

Excel_path = "\\now_excel\\xlsx"
save_path = 'now_excel'


def create_total_xlsx():
    if not os.path.exists(save_path + '\\' + '总表.xlsx'):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
    else:
        os.remove(save_path + '\\' + '总表.xlsx')
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Sheet1'
    for i in range(1, 19):
        data = [['第'+str(i)+'周', '星期一', '星期二', '星期三', '星期四', '星期五', '星期六', '星期日']]
        for row in data:
            sheet.append(row)
        for j in range(0, 6):
            data = [[j + 1]]
            for row in data:
                sheet.append(row)
    workbook.save(save_path + '\\' + '总表.xlsx')


def find_letter_position(letter):
    letter_positions = {}
    for index, char in enumerate(letter):
        # 检查是否为字母
        if char.isalpha():
            # 如果是字母，记录其位置
            # letter_positions[char] = letter_positions.get(char, []) + [index + 1]  # 索引从1开始
            return index


def xldate_as_datetime(xldate, datemode=0):
    if datemode not in (0, 1):
        raise Exception(datemode)
    if xldate == 0.00:
        return datetime.time(0, 0, 0)
    if xldate < 0.00:
        raise Exception(xldate)
    xldays = int(xldate)
    frac = xldate - xldays
    seconds = int(round(frac * 86400.0))
    assert 0 <= seconds <= 86400
    if seconds == 86400:
        seconds = 0
        xldays += 1
    # if xldays >= _XLDAYS_TOO_LARGE[datemode]:
    #    raise XLDateTooLarge(xldate)
    if xldays == 0:
        # second = seconds % 60; minutes = seconds // 60
        minutes, second = divmod(seconds, 60)
        # minute = minutes % 60; hour    = minutes // 60
        hour, minute = divmod(minutes, 60)
        return datetime.time(hour, minute, second)
    if xldays < 61 and datemode == 0:
        raise Exception(xldate)
    return (
            datetime.datetime.fromordinal(xldays + 693594 + 1462 * datemode)
            + datetime.timedelta(seconds=seconds)
    )


def add_data():
    retval = os.getcwd()
    retval = retval + Excel_path
    all_files = os.listdir(retval)

    save_workbook = openpyxl.load_workbook(os.getcwd() + '\\' + save_path + '\\' + '总表.xlsx')
    save_sheet = save_workbook['Sheet1']
    # 仅保留Excel文件
    excel_files = [file for file in all_files if file.endswith('.xlsx') or file.endswith('.xls')]

    for excel_file in excel_files:
        name = excel_file.split('.')[0]
        file_path = os.path.join(retval, excel_file)
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook['Sheet1']
        h = sheet.max_row
        w = sheet.max_column
        head_num = 0

        ss = sheet.cell(row=4, column=4).value
        print(ss)
        for i in range(1, h):
            if sheet.cell(row=i, column=1).value == '星期一':
                head_num = i
                break
        class_names = []
        teacher_names = []
        for i in range(3, head_num):
            class_names.append(sheet.cell(row=i, column=4).value)
        for i in range(head_num + 1, h):
            teacher_name = sheet.cell(row=i, column=8).value
            if teacher_name is None:
                continue
            teacher_names.append(teacher_name)
        if len(teacher_names) != len(class_names):
            print(name + '课程数量不对')
            return
        for i in range(head_num + 1, h):
            for j in range(1, 8):
                data = sheet.cell(row=i, column=j).value
                if data is None:
                    continue
                if isinstance(data, int):
                    continue

                letter_index = find_letter_position(data)
                head_data = data[: letter_index]
                if head_data == data:
                    continue
                print(head_data)
                print(data)
                if head_data == "0102":
                    save_row = 1
                elif head_data == "0304":
                    save_row = 2
                elif head_data == "0506":
                    save_row = 3
                elif head_data == "0708":
                    save_row = 4
                elif head_data == "0910":
                    save_row = 5
                elif head_data == "12":
                    save_row = 6
                else:
                    save_row = 6
                nums = 0
                while 1:
                    teacher_data = sheet.cell(row=i + nums, column=8).value
                    if teacher_data is None:
                        nums = nums - 1
                        continue
                    num_1 = 0
                    for teacher in teacher_names:
                        if teacher is teacher_data:
                            teacher_num = num_1
                            break
                        num_1 = num_1 + 1
                    break
                week_data = sheet.cell(row=i + 1, column=j).value
                if isinstance(week_data, int):
                    time_data = xldate_as_datetime(week_data)
                    head_time = time_data.month
                    end_time = time_data.day
                else:
                    time_data = week_data
                    time = time_data.split('-')
                    head_time = time[0]
                    end_time = time[1]

                print(head_time)
                print(end_time)
                for cycle in range(int(head_time) - 1, int(end_time)):
                    old = save_sheet.cell(row=save_row + 1 + cycle * 7, column=j + 1).value
                    if old == None:
                        old = ''
                    new = name + class_names[teacher_num]+" 忙碌\n"

                    combined_content = f"{old} {new}"
                    save_sheet.cell(save_row + 1 + cycle * 7, j + 1, combined_content)
    save_workbook.save(save_path + '\\' + '总表.xlsx')


def main(name):
    create_total_xlsx()
    add_data()


if __name__ == '__main__':
    main('PyCharm')
